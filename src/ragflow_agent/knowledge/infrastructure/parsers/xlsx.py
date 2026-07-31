"""Independent XLSX parser with formula and workbook resource controls."""

from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from ragflow_agent.knowledge.domain.chunk import (
    BlockKind,
    ParsedBlock,
    ParseWarning,
    TableMetadata,
)
from ragflow_agent.knowledge.domain.errors import KnowledgeConflictError
from ragflow_agent.knowledge.infrastructure.parsers.security import validate_zip_package
from ragflow_agent.knowledge.ports.parsing import (
    ParsedPayload,
    ParserCapability,
    ParseRequest,
)


class XlsxBinaryParser:
    """Parse non-empty worksheet regions without evaluating formulas or macros."""

    capability = ParserCapability(
        parser_id="independent-xlsx",
        parser_version="2",
        media_types=frozenset(
            {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        ),
        extensions=frozenset({".xlsx"}),
        default_chunk_strategy="table",
    )

    def __init__(
        self,
        *,
        max_entries: int,
        max_uncompressed_bytes: int,
        max_compression_ratio: float,
        max_sheets: int,
        max_rows_per_sheet: int,
        max_nonempty_cells: int,
    ) -> None:
        self._max_entries = max_entries
        self._max_uncompressed_bytes = max_uncompressed_bytes
        self._max_compression_ratio = max_compression_ratio
        self._max_sheets = max_sheets
        self._max_rows_per_sheet = max_rows_per_sheet
        self._max_nonempty_cells = max_nonempty_cells

    def parse_bytes(
        self,
        payload: bytes,
        request: ParseRequest,
        *,
        ocr_language: str,
    ) -> ParsedPayload:
        del request, ocr_language
        validate_zip_package(
            payload,
            max_entries=self._max_entries,
            max_uncompressed_bytes=self._max_uncompressed_bytes,
            max_compression_ratio=self._max_compression_ratio,
        )
        try:
            workbook = load_workbook(
                BytesIO(payload),
                read_only=False,
                data_only=False,
                keep_links=False,
            )
        except Exception as error:
            raise KnowledgeConflictError(
                "XLSX workbook is invalid",
                error_code="parser_xlsx_invalid",
            ) from error
        try:
            return self._parse_workbook(workbook)
        finally:
            workbook.close()

    def _parse_workbook(self, workbook: Workbook) -> ParsedPayload:
        if len(workbook.worksheets) > self._max_sheets:
            raise KnowledgeConflictError(
                "XLSX workbook has too many sheets",
                error_code="parser_resource_limit",
                details={"resource": "xlsx_sheets"},
            )
        blocks: list[ParsedBlock] = []
        warnings: list[ParseWarning] = []
        nonempty_cells = 0
        for worksheet in workbook.worksheets:
            if worksheet.max_row > self._max_rows_per_sheet:
                raise KnowledgeConflictError(
                    "XLSX worksheet exceeds the row limit",
                    error_code="parser_resource_limit",
                    details={
                        "resource": "xlsx_rows",
                        "sheet": worksheet.title,
                    },
                )
            rows: list[list[str]] = []
            formula_count = 0
            for cells in worksheet.iter_rows():
                row: list[str] = []
                for cell in cells:
                    value = cell.value
                    if value is not None:
                        nonempty_cells += 1
                        if nonempty_cells > self._max_nonempty_cells:
                            raise KnowledgeConflictError(
                                "XLSX workbook exceeds the non-empty cell limit",
                                error_code="parser_resource_limit",
                                details={"resource": "xlsx_nonempty_cells"},
                            )
                    if cell.data_type == "f":
                        formula_count += 1
                    row.append(self._render(value))
                while row and not row[-1]:
                    row.pop()
                if any(row):
                    rows.append(row)
            if not rows:
                continue
            blocks.append(
                ParsedBlock(
                    id=f"xlsx_{len(blocks)}",
                    kind=BlockKind.TABLE,
                    order=len(blocks),
                    text="\n".join("\t".join(row) for row in rows),
                    heading_path=(worksheet.title,),
                    table=TableMetadata(
                        rows=len(rows),
                        columns=max(len(row) for row in rows),
                        has_header=True,
                    ),
                )
            )
            if formula_count:
                warnings.append(
                    ParseWarning(
                        code="xlsx_formula_preserved_not_evaluated",
                        message=(
                            f"{formula_count} formula cells in {worksheet.title} "
                            "were preserved as text"
                        ),
                    )
                )
            if worksheet.merged_cells.ranges:
                warnings.append(
                    ParseWarning(
                        code="xlsx_merged_cells_present",
                        message=(
                            f"{len(worksheet.merged_cells.ranges)} merged ranges "
                            f"are present in {worksheet.title}"
                        ),
                    )
                )
        return ParsedPayload(blocks=tuple(blocks), warnings=tuple(warnings))

    @staticmethod
    def _render(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, (datetime, date, time)):
            return value.isoformat()
        if isinstance(value, bool):
            return "true" if value else "false"
        return str(value).strip()
